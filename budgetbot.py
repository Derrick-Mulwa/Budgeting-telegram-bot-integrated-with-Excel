import time
import asyncio
from datetime import datetime
import json
from telegram import *
from telegram.ext import *
from openpyxl import load_workbook
import random
import pandas as pd
import dataframe_image as dfi
import os
from openpyxl.styles import PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule
from openpyxl.formatting.rule import DataBar, FormatObject

# Bot details
TOKEN = '6795785508:AAH5AGjkJQj30Elsl8PuC_YC6jN1pDkM?4g'
BOT_USERNAME = '@BajceticBot'
MY_CHAT_ID = '1486454053'


# create a list to mark progress
ongoingProcess = [False, 'ProcessID', 'Process Description', 'Message to send']

# create a list for new items to buy later
# NewItem = ['ItemName', Itemprice]
NewItem = ['ItemName', 0]

# create a list for new poject to start
# NewProject = ['ProjectName', price]
NewProject = ['ProjectName', 0]

# Create a list for viewing month's budget
viewamonthsbudget_data = ['months_name_list', 'worksheets', 'months_budget_tables_list', 'Message ID']

# Create a list for adding an item to a month's budget
additemtoamonthsbudget_data = ['Months name', 'worksheets', 'months_budget_table','workbook', 'message_id']
item_to_budget_details = ['item name', 0]


# Create a function that will load the Excel file and create variable names for all the tables and worksheets
def load_excel():
    # Load the excel file

    wb = load_workbook(filename="BajeticAutomated.xlsx")

    # Create variables for all the worksheets and all the tables in each worksheet
    October = wb['October']
    OctoberBudget = October.tables['OctoberBudget']
    OctoberTotals = October.tables['OctoberTotals']
    OctoberSummary = October.tables['OctoberSummary']
    OctoberUnbudgeted = October.tables['OctoberUnbudgeted']

    october = [October, OctoberBudget, OctoberTotals, OctoberSummary, OctoberUnbudgeted]


    November = wb['November']
    NovemberBudget = November.tables['NovemberBudget']
    NovemberTotals = November.tables['NovemberTotals']
    NovemberSummary = November.tables['NovemberSummary']
    NovemberUnbudgeted = November.tables['NovemberUnbudgeted']

    november = [November, NovemberBudget, NovemberTotals,  NovemberSummary, NovemberUnbudgeted]

    December = wb['December']
    DecemberBudget = December.tables['DecemberBudget']
    DecemberTotals = December.tables['DecemberTotals']
    DecemberSummary = December.tables['DecemberSummary']
    DecemberUnbudgeted = December.tables['DecemberUnbudgeted']

    december = [December, DecemberBudget, DecemberTotals, DecemberSummary, DecemberUnbudgeted]

    LongTerm = wb['LongTerm']
    LongtermProjects = LongTerm.tables['LongtermProjects']

    longterm = [LongTerm, LongtermProjects]

    Items = wb['Items']
    ItemsTable = Items.tables['Items']

    items = [Items, ItemsTable]

    return [october, november, december, longterm, items, wb]



async def command_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print('Initialized start command.')
    await update.message.reply_text(f'Hello. Welcome to {BOT_USERNAME}. I am @Dev_orred personal budgeting bot.'
                                    f'\n\n\nContact him here: https://t.me/Dev_orred')


async def command_newitem(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess):
    # Check if there is an ongoing process. If not, get the data for the new item and add to the excel db
    message_id = update.message.message_id
    print('Intiated Add New Item command')
    if ongoingProcess[0] is False:
        ongoingProcess[0] = True
        ongoingProcess[1] = 'AddNewItem:Step1'
        ongoingProcess[2] = 'Adding a new Item'
        ongoingProcess[3] = 'You have an ongoing process: Adding a new Item. \n\nSend "Cancel Ongoing Process" to terminate the process'

        await update.message.reply_text('Enter the name of the new item', reply_to_message_id=message_id)

    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)



async def command_newproject(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess):
    # Check if there is an ongoing process. If not, get the data for the new project and add to the Excel db
    message_id = update.message.message_id
    print('Intiated Add New Project command')
    if ongoingProcess[0] is False:
        ongoingProcess[0] = True
        ongoingProcess[1] = 'AddNewProject:Step1'
        ongoingProcess[2] = 'Adding a new Project'
        ongoingProcess[3] = 'You have an ongoing process: Adding a new Project. \n\nSend "Cancel Ongoing Process" to terminate the process'

        await update.message.reply_text('Enter the name of the new project', reply_to_message_id=message_id)

    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)


# A command to send the items table.

async def command_viewitems(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess):
    # Check if there is an ongoing process. If not, get the data for the new project and add to the Excel db
    message_id = update.message.message_id
    if ongoingProcess[0] is False:
        print('Viewing Items')

        await update.message.reply_text(text= f'Processing...', reply_to_message_id=message_id)

        excel = load_excel()

        # Create variables table for the item's table and sheet for the item's sheet
        sheet = excel[4][0]
        table = excel[4][1]

        table_range = table.ref

        table_head = sheet[table_range][0]
        table_data = sheet[table_range][1:]

        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}

        for row in table_data:
            row_val = [cell.value for cell in row]
            for key, val in zip(columns, row_val):
                data[key].append(val)

        df = pd.DataFrame(data=data, columns=columns)

        df.index = range(1, len(df) + 1)
        df = df.rename_axis('Number', axis='index')

        df2 = df.copy()
        df = df.reset_index()

        # # Format the output text to be more readable
        data = [['Number', 'Item', 'Price']]
        data += [list(i) for i in df.values]

        output = ""
        for num, item, price in data:
            output += f"{num}. {item} - {price}\n"

        # PICTURIZE

        df = df2.copy()

        df['Price'] = df['Price'].apply(lambda x: f'{x:,.2f}')

        # Create a Styler object and set right alignment for numeric columns, and left for number and item
        styled_df = df.style.map(lambda x: 'text-align: left;', subset='Item').map(lambda x: 'text-align: right;',
                                                                                   subset='Price')

        # Render the styled dataframe
        dfi.export(styled_df, "dataframe_table.png")

        path = 'dataframe_table.png'
        await context.bot.deleteMessage(chat_id=update.message.chat_id, message_id=message_id+1)
        await update.message.reply_photo(photo= path, caption= f'These are the items in your list', reply_to_message_id=message_id)

        os.remove(path)


    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)


# Create a command handler to fetch and show projects list
async def command_viewprojects(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess):
    # Check if there is an ongoing process. If not, get the data for the new project and add to the Excel db
    message_id = update.message.message_id
    if ongoingProcess[0] is False:
        print('Viewing projects')

        await update.message.reply_text(text= f'Fetching your projects. Please wait...', reply_to_message_id=message_id)

        excel = load_excel()

        # Create variables table for the item's table and sheet for the item's sheet
        sheet = excel[3][0]
        table = excel[3][1]

        table_range = table.ref

        table_head = sheet[table_range][0]
        table_data = sheet[table_range][1:]

        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}

        for row in table_data:
            row_val = [cell.value for cell in row]
            for key, val in zip(columns, row_val):
                data[key].append(val)

        df = pd.DataFrame(data=data, columns=columns)

        df.index = range(1, len(df) + 1)
        df = df.rename_axis('Number', axis='index')

        df['Amount'] = df['Amount'].apply(lambda x: f'{x:,.2f}')

        # Create a Styler object and set right alignment for numeric columns, and left for number and item
        styled_df = df.style.map(lambda x: 'text-align: left;', subset='Project').map(lambda x: 'text-align: right;',
                                                                                   subset='Amount')

        # Render the styled dataframe
        dfi.export(styled_df, "dataframe_table.png")

        path = 'dataframe_table.png'

        path = 'dataframe_table.png'
        await context.bot.deleteMessage(chat_id=update.message.chat_id, message_id=message_id+1)
        await update.message.reply_photo(photo= path, caption= f'These are the projects you have', reply_to_message_id=message_id)

        os.remove(path)

    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)

# CREATE A FEW FUNCTIONS FOR THE COMMAND BELOW: viewamonthsbudget
def percentage_completed(dataframe):
    data = dataframe.values

    per = []
    for i in range(len(data)):
        try:
            value = (data[i][4] ) * 100 / data[i][2]
        except:
            value = 0

        per.append(value)

    return per

def over_budget(dataframe):
    data = dataframe.values
    per = []

    for i in range(len(data)):
        if data[i][5] == 'YES':
            value = data[i][4] - data[i][2]
        else:
            value = 0

        per.append(value)

    return per

def amt_used(dataframe):
    data = dataframe.values
    amt = []
    for i in range(len(data)):
        try:
            value = int(data[i][4])
            if value < 0:
                value = 0
        except:
            value = 0

        amt.append(value)

    return amt

# Command view month's budget
async def command_viewamonthsbudget(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess, viewamonthsbudget_data = viewamonthsbudget_data):
    # Check if there is an ongoing process. If not, get the data for the new project and add to the Excel db
    message_id = update.message.message_id
    if ongoingProcess[0] is False:
        print('Viewing a months budget command initialised')

        ongoingProcess[0] = True
        ongoingProcess[1] = 'ViewAmonthsBuget:Step1'
        ongoingProcess[2] = 'Viewing a month\'s budget'
        ongoingProcess[3] = 'You have an ongoing process: Viewing a month\'s budget. \n\nSend "Cancel Ongoing Process" to terminate the process'


        excel = load_excel()
        months = excel[:-3]

        months_name = [i[0].title for i in months]
        worksheets = [i[0] for i in months]
        months_budget_table = [i[1] for i in months]

        viewamonthsbudget_data[0] = months_name
        viewamonthsbudget_data[1] = worksheets
        viewamonthsbudget_data[2] = months_budget_table
        viewamonthsbudget_data[3] = message_id

        buttons = []
        for month in months_name:
            button_data = [InlineKeyboardButton(month, callback_data=f'ViewaMonthsBudget:{month}')]
            buttons.append(button_data)

        await update.message.reply_text('Select a month to view the budget.', reply_to_message_id=message_id, reply_markup=InlineKeyboardMarkup(buttons))

    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)




async def command_addItemToAmonthsbudget(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess, additemtoamonthsbudget_data = additemtoamonthsbudget_data):
    message_id = update.message.message_id
    if ongoingProcess[0] is False:
        print('Adding item to a months budget command initialised')

        ongoingProcess[0] = True
        ongoingProcess[1] = 'AddItemToAmonthsBuget:Step1'
        ongoingProcess[2] = 'Adding item to a month\'s budget'
        ongoingProcess[
            3] = 'You have an ongoing process: Adding an item to a month\'s budget. \n\nSend "Cancel Ongoing Process" to terminate the process'

        excel = load_excel()
        months = excel[:-3]

        months_name = [i[0].title for i in months]
        worksheets = [i[0] for i in months]
        months_budget_table = [i[1] for i in months]

        additemtoamonthsbudget_data[0] = months_name
        additemtoamonthsbudget_data[1] = worksheets
        additemtoamonthsbudget_data[2] = months_budget_table
        additemtoamonthsbudget_data[3] = excel[5]
        additemtoamonthsbudget_data[4] = message_id


        buttons = []
        for month in months_name:
            button_data = [InlineKeyboardButton(month, callback_data=f'AddItemtoaMonthsBudget:{month}')]
            buttons.append(button_data)

        await update.message.reply_text('Select a month to add item to its budget.', reply_to_message_id=message_id,
                                        reply_markup=InlineKeyboardMarkup(buttons))

    else:
        await update.message.reply_text(f'{ongoingProcess[3]}', reply_to_message_id=message_id)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess,
                         NewItem = NewItem, NewProject = NewProject, viewamonthsbudget_data=viewamonthsbudget_data,
                         additemtoamonthsbudget_data = additemtoamonthsbudget_data,
                         item_to_budget_details=item_to_budget_details):

    message_text = update.message.text
    message_id = update.message.message_id

    # Cancel any ongoing process and reset it using the "Cancel Ongoing Process" keyword
    if message_text.lower() == 'cancel ongoing process':
        print('Cancelling any ongoing process')
        ongoingProcess[0] = False
        ongoingProcess[1] = 'ProcessID'
        ongoingProcess[2] = 'Process Description'
        ongoingProcess[3] = 'Message to send'

        NewItem[0] = 'ItemName'
        NewItem[1] = 0

        NewProject[0] = 'ProjectName'
        NewProject[1] = 0

        viewamonthsbudget_data[0] = 'months_name_list'
        viewamonthsbudget_data[1] = 'worksheets'
        viewamonthsbudget_data[2] = 'months_budget_tables_list'

        additemtoamonthsbudget_data[0] = 'months_name_list'
        additemtoamonthsbudget_data[1] = 'worksheets'
        additemtoamonthsbudget_data[2] = 'months_budget_tables_list'
        additemtoamonthsbudget_data[3] = 'workbook'
        additemtoamonthsbudget_data[4] = 'Message ID'

        item_to_budget_details[0] = 'item name'
        item_to_budget_details[1] = 0


        await update.message.reply_text(f'Any ongoing process has been reset successfully', reply_to_message_id=message_id)

    # ADDING A NEW ITEM
    # Handle the result of add new item: Step 1
    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddNewItem:Step1':
        item_name = message_text
        if item_name != '':
            ongoingProcess[1] = 'AddNewItem:Step2'
            ongoingProcess[3] = f'You have an ongoing process: Adding price for new item: {item_name}.\n\nSend "Cancel Ongoing Process" to terminate the process'
            NewItem[0] = item_name
            print('Requesting price for new Item')
            await update.message.reply_text(f'Enter the price for the new item: {item_name}', reply_to_message_id=message_id)

        else:
            await update.message.reply_text(f'Message cannot be empty. Please re-enter the new item\'s name.', reply_to_message_id=message_id)



    # Adding a new item after collecting all the details : Step 2
    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddNewItem:Step2':
        item_price = message_text
        try:
            item_price = int(item_price)
            if item_price <= 0:
                await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0",
                                                reply_to_message_id=message_id)

            else:
                try:
                    print('Adding a new Item')

                    NewItem[1] = item_price

                    # Write on Excel
                    # print(NewItem)

                    # Load the Excel
                    excel = load_excel()

                    Items_worksheet = excel[4][0]
                    Items_table = excel[4][1]

                    workbook = excel[5]

                    ref = Items_table.ref

                    last_row_number = ref[ref.index(":") + 1:][[index for index, char in enumerate(ref[ref.index(":") + 1:]) if char.isdigit()][0]:]
                    new_row_number = int(last_row_number)+1

                    Items_worksheet[f'A{new_row_number}'] = NewItem[0]
                    Items_worksheet[f'B{new_row_number}'] = NewItem[1]

                    table_start = ref[:ref.index(":") + 1]

                    new_table_end_column = ref[ref.index(":") + 1:][:[index for index, char in
                                                                      enumerate(ref[ref.index(":") + 1:]) if
                                                                      char.isdigit()][0]]
                    new_table_end_row = new_row_number
                    table_end = f'{new_table_end_column}{new_table_end_row}'

                    new_table_ref = f'{table_start}{table_end}'

                    Items_table.ref = new_table_ref

                    workbook.save("BajeticAutomated.xlsx")

                    await update.message.reply_text(text=f"New item added successfully!\nName: {NewItem[0]}\nPrice: {NewItem[1]}",
                                                    reply_to_message_id=message_id)

                    ongoingProcess[0] = False
                    ongoingProcess[1] = 'ProcessID'
                    ongoingProcess[2] = 'Process Description'
                    ongoingProcess[3] = 'Message to send'

                    NewItem[0] = 'ItemName'
                    NewItem[1] = 0
                except:
                    await update.message.reply_text(text="An unexpected error occurred. Please try again later. "
                                                         "Apologies for the inconvenience caused", reply_to_message_id=message_id)

        except:
            await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0", reply_to_message_id=message_id)




    # ADDING A NEW PROJECT
    # Handle the result of add new item: Step 1
    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddNewProject:Step1':
        project_name = message_text
        if project_name != '':
            ongoingProcess[1] = 'AddNewProject:Step2'
            ongoingProcess[
                3] = f'You have an ongoing process: Adding price for new project: {project_name}.\n\nSend "Cancel Ongoing Process" to terminate the process'
            NewProject[0] = project_name
            print('Requesting price for new Item')
            await update.message.reply_text(f'Enter the price for the new project: {project_name}',
                                            reply_to_message_id=message_id)

        else:
            await update.message.reply_text(f'Project name MUST NOT be empty. Please re-enter the new project\'s name.',
                                            reply_to_message_id=message_id)



    # Adding a new project after collecting all the details for the project: Step 2
    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddNewProject:Step2':
        project_price = message_text
        try:
            project_price = int(project_price)
            if project_price <= 0:
                await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0",
                                                reply_to_message_id=message_id)

            else:
                try:
                    print('Adding a new project')

                    NewProject[1] = project_price

                    # Write on Excel
                    # print(NewProject)

                    # Load the Excel
                    excel = load_excel()

                    Projects_worksheet = excel[3][0]
                    Projects_table = excel[3][1]

                    workbook = excel[5]

                    ref = Projects_table.ref

                    last_row_number = ref[ref.index(":") + 1:][
                                      [index for index, char in enumerate(ref[ref.index(":") + 1:]) if char.isdigit()][
                                          0]:]
                    new_row_number = int(last_row_number) + 1

                    Projects_worksheet[f'A{new_row_number}'] = NewProject[0]
                    Projects_worksheet[f'B{new_row_number}'] = NewProject[1]

                    table_start = ref[:ref.index(":") + 1]

                    new_table_end_column = ref[ref.index(":") + 1:][:[index for index, char in
                                                                      enumerate(ref[ref.index(":") + 1:]) if
                                                                      char.isdigit()][0]]
                    new_table_end_row = new_row_number
                    table_end = f'{new_table_end_column}{new_table_end_row}'

                    new_table_ref = f'{table_start}{table_end}'

                    Projects_table.ref = new_table_ref

                    workbook.save("BajeticAutomated.xlsx")

                    await update.message.reply_text(
                        text=f"New project added successfully!\nName: {NewProject[0]}\nPrice: {NewProject[1]}",
                        reply_to_message_id=message_id)

                    ongoingProcess[0] = False
                    ongoingProcess[1] = 'ProcessID'
                    ongoingProcess[2] = 'Process Description'
                    ongoingProcess[3] = 'Message to send'

                    NewProject[0] = 'ProjectName'
                    NewProject[1] = 0
                except:
                    await update.message.reply_text(text="An unexpected error occurred. Please try again later. \n"
                                                         "Apologies for the inconvenience caused",
                                                    reply_to_message_id=message_id)

        except:
            await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0",
                                            reply_to_message_id=message_id)


    # Handle the result of add new item to budget: Step 2 Getting item name
    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddItemToAmonthsBuget:Step2':
        item_name = message_text
        if item_name != '':
            ongoingProcess[1] = 'AddItemToAmonthsBuget:Step3'
            ongoingProcess[
                3] = f'You have an ongoing process: Adding price for item to month\'s budget: {item_name}.\n\n' \
                     f'Send "Cancel Ongoing Process" to terminate the process'
            item_to_budget_details[0] = item_name
            print('Requesting price for new Item')
            await update.message.reply_text(f'Enter the price for the item: {item_name}',
                                            reply_to_message_id=message_id)

        else:
            await update.message.reply_text(f'Item name MUST NOT be empty. Please re-enter the item\'s name.',
                                            reply_to_message_id=message_id)

    elif (ongoingProcess[0] is True) and ongoingProcess[1] == 'AddItemToAmonthsBuget:Step3':
        item_price = message_text
        try:
            item_price = int(item_price)
            item_to_budget_details[1] = item_price
            if item_price <= 0:
                await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0",
                                                reply_to_message_id=message_id)
            else:
                month = additemtoamonthsbudget_data[-2]
                months_name = additemtoamonthsbudget_data[0]
                worksheets = additemtoamonthsbudget_data[1]
                months_budget_table = additemtoamonthsbudget_data[2]
                workbook = additemtoamonthsbudget_data[3]
                message_id = additemtoamonthsbudget_data[4]

                index = months_name.index(month)
                Budget_table = months_budget_table[index]
                Budget_worksheet = worksheets[index]

                ref = Budget_table.ref

                _date = datetime.now().strftime('%d-%m-%Y')
                item_name = item_to_budget_details[0]
                item_price = item_to_budget_details[1]

                last_row_number = ref[ref.index(":") + 1:][
                                  [index for index, char in enumerate(ref[ref.index(":") + 1:]) if char.isdigit()][
                                      0]:]
                new_row_number = int(last_row_number) + 1

                print(new_row_number)

                Budget_worksheet[f'A{new_row_number}'] = _date
                Budget_worksheet[f'B{new_row_number}'] = item_name
                Budget_worksheet[f'C{new_row_number}'] = item_price
                Budget_worksheet[f'F{new_row_number}'] = 'NO'

                table_start = ref[:ref.index(":") + 1]

                new_table_end_column = ref[ref.index(":") + 1:][:[index for index, char in
                                                                  enumerate(ref[ref.index(":") + 1:]) if
                                                                  char.isdigit()][0]]
                new_table_end_row = new_row_number
                table_end = f'{new_table_end_column}{new_table_end_row}'

                new_table_ref = f'{table_start}{table_end}'

                Budget_table.ref = new_table_ref

                Budget_worksheet[f'D{new_row_number}'] = f'=E{new_row_number}*100/C{new_row_number}'
                Budget_worksheet[
                    f'G{new_row_number}'] = f'=IF(F{new_row_number}="YES",E{new_row_number}-C{new_row_number},0)'

                # Color formatting
                # Clear all prior formatting
                all_current_formating_rules = list(Budget_worksheet.conditional_formatting._cf_rules)
                column_C_and_E_formatting_rules = [i for i in all_current_formating_rules if
                                                   'C1' in str(i) or 'C2' in str(i)
                                                   or 'E1' in str(i) or 'E2' in str(i) or 'F1' in str(i) or 'F2' in str(
                                                       i)
                                                   or 'G1' in str(i) or 'G2' in str(i)]

                for i in column_C_and_E_formatting_rules:
                    del Budget_worksheet.conditional_formatting._cf_rules[i]

                # Create a color scale rule
                color_scale_rule = ColorScaleRule(
                    start_type='min',
                    start_color="12B61D",  # Green
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFF000',  # Yellow
                    end_type='max',
                    end_color="E81313"  # Red
                )

                # Apply the color scale rule to desired columns
                Budget_worksheet.conditional_formatting.add(f'C2:C{new_row_number}', color_scale_rule)
                Budget_worksheet.conditional_formatting.add(f'E2:E{new_row_number}', color_scale_rule)
                Budget_worksheet.conditional_formatting.add(f'G2:G{new_row_number}', color_scale_rule)

                # Create Data bar rule (To be used later)
                # first = FormatObject(type='min')
                # second = FormatObject(type='max')
                # data_bar = DataBar(cfvo=[first, second], color="00FF00", showValue=True, minLength=None, maxLength=None)
                # # assign the data bar to a rule
                # rule = Rule(type='dataBar', dataBar=data_bar)
                # Apply t cell range
                # Budget_worksheet.conditional_formatting.add(f'C1:C{new_row_number}', rule)

                # Create a fill rule for ompleted(Yes/no) column
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                rule1 = CellIsRule(operator="equal", formula=['"NO"'], fill=red_fill)

                green_fill = PatternFill(start_color="12B61D", end_color="12B61D", fill_type="solid")
                rule2 = CellIsRule(operator="equal", formula=['"YES"'], fill=green_fill)

                # Add the conditional formatting rule to the worksheet
                Budget_worksheet.conditional_formatting.add(f'F2:F{new_row_number}', rule1)
                Budget_worksheet.conditional_formatting.add(f'F2:F{new_row_number}', rule2)

                workbook.save("BajeticAutomated.xlsx")

                await update.message.reply_text(text=f"{item_name} successfully added to {month}'s budget\n\n"
                                                     f"Item : {item_name}\n"
                                                     f"Price: {item_price}\n"
                                                     f"Month: {month}",
                                                reply_to_message_id=message_id)

                additemtoamonthsbudget_data[0] = 'months_name_list'
                additemtoamonthsbudget_data[1] = 'worksheets'
                additemtoamonthsbudget_data[2] = 'months_budget_tables_list'
                additemtoamonthsbudget_data[3] = 'workbook'
                additemtoamonthsbudget_data[4] = 'Message ID'

                item_to_budget_details[0] = 'item name'
                item_to_budget_details[1] = 0

                ongoingProcess[0] = False
                ongoingProcess[1] = 'ProcessID'
                ongoingProcess[2] = 'Process Description'
                ongoingProcess[3] = 'Message to send'


        except:
            await update.message.reply_text(text="Invalid Price! Please enter a number greater than 0",
                                            reply_to_message_id=message_id)



    # Reply if the message has no meaning or use
    else:
        # media = [[media_id, media type]]
        # Send a media to explain that the message sent has no meaning and use
        media = [['BAACAgQAAxkBAAIBCWVDo1DWBGdPmg38NisgQFPFmHB9AAIWEQAClx0ZUrxoX5A-uFgsMwQ', 'VIDEO'],
                 ['BAACAgQAAxkBAAIBDGVDo_GQo9yuo885SJSwuZ9MmDEhAAIXEQAClx0ZUorMcdDhyEbkMwQ', 'VIDEO'],
                 ['AgACAgQAAxkBAAIBBWVDoR0YPunb7Tayo7fAQMNCO_IDAAJhvzEblx0ZUqAnFxskHCDcAQADAgADcwADMwQ', 'PHOTO']]

        media_to_send = random.choice(media)

        if media_to_send[1] == 'VIDEO':
            await update.message.reply_video(caption="I don't know what you want me to do with this information my G.\n\n"
                                                     "Please start a process and then I'll be of help",
                                             reply_to_message_id=message_id, video=media_to_send[0])
        if media_to_send[1] == 'PHOTO':
            await update.message.reply_photo(caption="I don't know what you want me to do with this information my G.\n\n"
                                                     "Please start a process and then I'll be of help",
                                             reply_to_message_id=message_id, photo = media_to_send[0])




# Create a querry handler
async def callbackhandler(update: Update, context: CallbackContext, viewamonthsbudget_data = viewamonthsbudget_data, additemtoamonthsbudget_data = additemtoamonthsbudget_data):
    querry = update.callback_query
    await update.callback_query.answer()
    chat = querry.message.chat_id
    data = querry.data
    querry_message_id = querry.message.message_id

    if data[:18] == 'ViewaMonthsBudget:':
        month = data[18:]
        print(f'Sending {month}\'s budget')


        await context.bot.editMessageText(chat_id=chat, message_id=querry_message_id, text='Processing...')

        months_name = viewamonthsbudget_data[0]
        worksheets = viewamonthsbudget_data[1]
        months_budget_table = viewamonthsbudget_data[2]
        message_id = viewamonthsbudget_data[3]

        index = months_name.index(month)
        table = months_budget_table[index]
        sheet = worksheets[index]

        table_range = table.ref

        table_head = sheet[table_range][0]
        table_data = sheet[table_range][1:]

        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}

        for row in table_data:
            row_val = [cell.value for cell in row]
            for key, val in zip(columns, row_val):
                data[key].append(val)

        df = pd.DataFrame(data=data, columns=columns)

        df.index = range(1, len(df) + 1)
        df = df.rename_axis('Number', axis='index')
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)

        df['Amount used'] = amt_used(df)
        df['Percentage Completed'] = percentage_completed(df)
        df['Over budget'] = over_budget(df)

        # Format the numbers
        df['Price'] = df['Price'].apply(lambda x: f'{x:,.2f}')
        df['Amount used'] = df['Amount used'].apply(lambda x: f'{x:,.2f}')
        df['Over budget'] = df['Over budget'].apply(lambda x: f'{x:,.2f}')
        df['Percentage Completed'] = df['Percentage Completed'].apply(lambda x: f'{x:,.2f}')

        styled_df = df.style.map(lambda x: 'text-align: left;', subset='Item').map(lambda x: 'text-align: right;',
                                                                                   subset='Price')

        # Render the styled dataframe
        dfi.export(styled_df, "dataframe_table.png")

        await context.bot.deleteMessage(chat_id=chat, message_id=querry_message_id)

        await context.bot.send_photo(photo="dataframe_table.png", caption=f"Here is your {month}'s budget.",
                                         reply_to_message_id=message_id, chat_id=chat)


        os.remove("dataframe_table.png")
        ongoingProcess[0] = False
        ongoingProcess[1] = 'ProcessID'
        ongoingProcess[2] = 'Process Description'
        ongoingProcess[3] = 'Message to send'


        viewamonthsbudget_data[0] = 'months_name_list'
        viewamonthsbudget_data[1] = 'worksheets'
        viewamonthsbudget_data[2] = 'months_budget_tables_list'




    elif data[:22]  == 'AddItemtoaMonthsBudget':
        month = data[23:]
        print(f'Adding item to budget for {month}: Requesting item name.')
        additemtoamonthsbudget_data += [month, querry_message_id]

        ongoingProcess[1] = 'AddItemToAmonthsBuget:Step2'
        ongoingProcess[2] = 'Adding item name to a month\'s budget'


        await context.bot.editMessageText(chat_id=chat, message_id=querry_message_id,
                                          text=f'Enter the name of the item to add to {month}\'s budget')






# CREATE A COMMAND TO RESET AND CANCEL ANY ONGOING PROCESS
async def command_cancelongoingprocess(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess, NewItem = NewItem, viewamonthsbudget_data=viewamonthsbudget_data):
    print('Cancelling any ongoing process')
    message_id = update.message.message_id

    ongoingProcess[0] = False
    ongoingProcess[1] = 'ProcessID'
    ongoingProcess[2] = 'Process Description'
    ongoingProcess[3] = 'Message to send'

    NewItem[0] = 'ItemName'
    NewItem[1] = 0

    NewProject[0] = 'ProjectName'
    NewProject[1] = 0

    viewamonthsbudget_data[0] = 'months_name_list'
    viewamonthsbudget_data[1] = 'worksheets'
    viewamonthsbudget_data[2] = 'months_budget_tables_list'

    additemtoamonthsbudget_data[0] = 'months_name_list'
    additemtoamonthsbudget_data[1] = 'worksheets'
    additemtoamonthsbudget_data[2] = 'months_budget_tables_list'
    additemtoamonthsbudget_data[3] = 'workbook'
    additemtoamonthsbudget_data[4] = 'Message ID'

    item_to_budget_details[0] = 'item name'
    item_to_budget_details[1] = 0

    await update.message.reply_text(f'Any ongoing process has been reset successfully', reply_to_message_id=message_id)



# CREATE A COMMAND TO RESET AND CANCEL ANY ONGOING PROCESS
async def command_list(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess, NewItem = NewItem, viewamonthsbudget_data=viewamonthsbudget_data):
    print('sending command list')
    message_id = update.message.message_id
    # _italic_[link](http: // example.com)

    command_list_text = f'You can control me by using these commands:\n\n' \
                        f'*Viewing*\n' \
                        f'/viewmyitems : View all the items you want to buy later in the items table\n' \
                        f'/viewmyprojects : View all the projects you have in your long term projects table\n' \
                        f'/viewamonthsbudget : View the budget of a specific month\n\n' \
                        f'' \
                        f'*Adding items*\n' \
                        f'/newitem : Add a new items to the items you want to buy later list\n' \
                        f'/newproject : Add a new project to the long term projects table\n' \
                        f'/newitemtobudget: Add a new item to a specific months budget\. Include it into the budget\n\n' \
                        f'' \
                        f'*Cancel ongoing processes*\n' \
                        f'/cancelongoingprocess : Cancel all ongoing processes'

    # Send the message with mixed formatting
    await update.message.reply_text(text = command_list_text, parse_mode='MarkdownV2',  reply_to_message_id=message_id)

async def handle_VID(update: Update, context: ContextTypes.DEFAULT_TYPE, ongoingProcess = ongoingProcess, NewItem = NewItem):
    print(update)



if __name__ == '__main__':
    print('Starting bot')
    app = Application.builder().token(TOKEN).build()

    # add command handlers for bot commands
    app.add_handler(CommandHandler('start', command_start))
    app.add_handler(CommandHandler('commands', command_list))

    app.add_handler(CommandHandler('newitem', command_newitem))
    app.add_handler(CommandHandler('cancelongoingprocess', command_cancelongoingprocess))
    app.add_handler(CommandHandler('newproject', command_newproject))
    app.add_handler(CommandHandler('viewmyitems', command_viewitems))
    app.add_handler(CommandHandler('viewmyprojects', command_viewprojects))
    app.add_handler(CommandHandler('viewamonthsbudget', command_viewamonthsbudget))
    app.add_handler(CommandHandler('newitemtobudget', command_addItemToAmonthsbudget))




    # Querry handler
    app.add_handler(CallbackQueryHandler(callbackhandler))

    # message Handler
    app.add_handler(MessageHandler(filters.TEXT, handle_message))
    app.add_handler(MessageHandler(filters.VIDEO, handle_VID))
    app.add_handler(MessageHandler(filters.PHOTO, handle_VID))



    # Add error handler
    app.add_error_handler(error)
    print('Polling...')
    app.run_polling(poll_interval=1)