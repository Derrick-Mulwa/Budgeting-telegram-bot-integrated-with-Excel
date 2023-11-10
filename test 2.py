from budgetbot import load_excel
from datetime import datetime
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import DataBar, FormatObject


additemtoamonthsbudget_data = ['months_name_list', 'worksheets', 'months_budget_tables_list','workbook', 'Message ID']
excel = load_excel()
months = excel[:-3]

months_name = [i[0].title for i in months]
worksheets = [i[0] for i in months]
months_budget_table = [i[1] for i in months]

additemtoamonthsbudget_data[0] = months_name
additemtoamonthsbudget_data[1] = worksheets
additemtoamonthsbudget_data[2] = months_budget_table
additemtoamonthsbudget_data[3] = excel[5]
additemtoamonthsbudget_data[4] = 'message_id'


additemtoamonthsbudget_data += ['October', 939]


month = additemtoamonthsbudget_data[-2]
print(month)
months_name = additemtoamonthsbudget_data[0]
worksheets = additemtoamonthsbudget_data[1]
months_budget_table = additemtoamonthsbudget_data[2]
workbook = additemtoamonthsbudget_data[3]
message_id = additemtoamonthsbudget_data[4]

index = months_name.index(month)
Budget_table = months_budget_table[index]
Budget_worksheet = worksheets[index]

ref = Budget_table.ref



item_to_budget_details = ['GR 131', 6605]

_date = datetime.now().strftime('%d-%m-%Y')
item_name = item_to_budget_details[0]
item_price = item_to_budget_details[1]

last_row_number = ref[ref.index(":") + 1:][
                                      [index for index, char in enumerate(ref[ref.index(":") + 1:]) if char.isdigit()][
                                          0]:]
new_row_number = int(last_row_number) + 1

print(new_row_number)

Budget_worksheet[f'A{new_row_number}'] = _date
Budget_worksheet[f'B{new_row_number}'] = item_name
Budget_worksheet[f'C{new_row_number}'] = item_price
Budget_worksheet[f'F{new_row_number}'] = 'YES'


table_start = ref[:ref.index(":") + 1]

new_table_end_column = ref[ref.index(":") + 1:][:[index for index, char in
                                                  enumerate(ref[ref.index(":") + 1:]) if
                                                  char.isdigit()][0]]
new_table_end_row = new_row_number
table_end = f'{new_table_end_column}{new_table_end_row}'

new_table_ref = f'{table_start}{table_end}'

Budget_table.ref = new_table_ref

Budget_worksheet[f'D{new_row_number}'] = f'=E{new_row_number}*100/C{new_row_number}'
Budget_worksheet[f'G{new_row_number}'] = f'=IF(F{new_row_number}="YES",E{new_row_number}-C{new_row_number},0)'


# Color formatting
# Clear all prior formatting
all_current_formating_rules = list(Budget_worksheet.conditional_formatting._cf_rules)
column_C_and_E_formatting_rules = [i for i in all_current_formating_rules if 'C1' in str(i) or 'C2' in str(i)
                                   or 'E1' in str(i) or 'E2' in str(i) or 'F1' in str(i) or 'F2' in str(i)
                                   or 'G1' in str(i) or 'G2' in str(i)]

for i in column_C_and_E_formatting_rules:
    del Budget_worksheet.conditional_formatting._cf_rules[i]


# Create a color scale rule
color_scale_rule = ColorScaleRule(
                    start_type='min',
                    start_color="12B61D",  # Green
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFF000',  # Yellow
                    end_type='max',
                    end_color = "E81313"  # Red
)


# Apply the color scale rule to desired columns
Budget_worksheet.conditional_formatting.add(f'C2:C{new_row_number}', color_scale_rule)
Budget_worksheet.conditional_formatting.add(f'E2:E{new_row_number}', color_scale_rule)
Budget_worksheet.conditional_formatting.add(f'G2:G{new_row_number}', color_scale_rule)


# Create Data bar rule (To be used later)
# first = FormatObject(type='min')
# second = FormatObject(type='max')
# data_bar = DataBar(cfvo=[first, second], color="00FF00", showValue=True, minLength=None, maxLength=None)
# # assign the data bar to a rule
# rule = Rule(type='dataBar', dataBar=data_bar)
# Apply t cell range
# Budget_worksheet.conditional_formatting.add(f'C1:C{new_row_number}', rule)






# Create a fill rule for ompleted(Yes/no) column
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
rule1 = CellIsRule(operator="equal", formula=['"NO"'], fill=red_fill)

green_fill = PatternFill(start_color="12B61D", end_color="12B61D", fill_type="solid")
rule2 = CellIsRule(operator="equal", formula=['"YES"'], fill=green_fill)

# Add the conditional formatting rule to the worksheet
Budget_worksheet.conditional_formatting.add(f'F2:F{new_row_number}', rule1)
Budget_worksheet.conditional_formatting.add(f'F2:F{new_row_number}', rule2)


workbook.save("BajeticAutomated.xlsx")


